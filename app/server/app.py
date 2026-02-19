"""Flask app, route registration, and static serving."""

import atexit
import functools
import json
import os
import queue
import sys

from flask import Flask, Response, g, jsonify, request, send_file, send_from_directory, session, stream_with_context

from .auth import (
    validate_credentials, get_active_sessions, set_active_session,
    clear_active_session, check_session_conflict, get_or_create_secret_key,
)
from .database import get_connection, init_db, row_to_dict, rows_to_dicts
from .events import subscribe, unsubscribe
from .importer import import_folder as do_import, IMAGE_EXTENSIONS
from .watcher import start_watcher, stop_watcher

LIBRARY_ROOT = os.getcwd()
APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # app/


def _settings_path():
    return os.path.join(LIBRARY_ROOT, '.library', 'settings.json')


def _load_settings():
    path = _settings_path()
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_settings(data):
    path = _settings_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w') as f:
        json.dump(data, f, indent=4)


IMAGE_ROOT = _load_settings().get('image_root') or LIBRARY_ROOT


def _clear_all_marks():
    """Reset all marked images to active on shutdown."""
    try:
        conn = get_connection()
        conn.execute(
            "UPDATE images SET status = 'active', modified_at = CURRENT_TIMESTAMP "
            "WHERE status IN ('marked_delete', 'marked_ocr')"
        )
        conn.commit()
        conn.close()
    except Exception:
        pass


def _migrate_thumbnail_dirs():
    """Migrate thumbnail directories from name-based to ID-based.

    Old format: .library/thumbnails/<folder_name>/
    New format: .library/thumbnails/<folder_id>/
    Also updates thumbnail_path in the images table.
    """
    try:
        thumb_root = os.path.join(LIBRARY_ROOT, '.library', 'thumbnails')
        if not os.path.isdir(thumb_root):
            return

        conn = get_connection()
        folders = conn.execute("SELECT id, name FROM folders").fetchall()

        for f in folders:
            fid = str(f['id'])
            fname = f['name']
            old_dir = os.path.join(thumb_root, fname)
            new_dir = os.path.join(thumb_root, fid)

            # Already migrated or new-style
            if os.path.isdir(new_dir):
                continue

            # Old name-based dir exists — rename it
            if os.path.isdir(old_dir):
                os.rename(old_dir, new_dir)
                # Update thumbnail_path in images: "FolderName/file.jpg" → "id/file.jpg"
                conn.execute(
                    "UPDATE images SET thumbnail_path = REPLACE(thumbnail_path, ?, ?) WHERE folder_id = ?",
                    (fname + '/', fid + '/', f['id'])
                )

        conn.commit()
        conn.close()
    except Exception:
        pass


def create_app():
    """Create and configure the Flask application."""
    app = Flask(
        __name__,
        static_folder=os.path.join(APP_DIR, 'static'),
        static_url_path='/static',
    )

    app.secret_key = get_or_create_secret_key()

    init_db()
    _clear_all_marks()  # Clear stale marks from previous session
    _migrate_thumbnail_dirs()  # Move name-based thumb dirs to ID-based
    atexit.register(_clear_all_marks)
    start_watcher(IMAGE_ROOT)
    atexit.register(stop_watcher)

    # ── Auth middleware ──────────────────────────────────────────────

    @app.before_request
    def _set_user_role():
        """Set g.user_role from session cookie + IP validation."""
        role = session.get('role')
        ip = request.remote_addr
        if role and role in ('data_entry', 'warehouse'):
            # Validate the session IP still matches
            sessions = get_active_sessions()
            active = sessions.get(role)
            if active and active['ip'] == ip:
                g.user_role = role
                g.username = session.get('username', '')
            else:
                # Session invalid (IP mismatch or cleared)
                session.clear()
                g.user_role = 'viewer'
                g.username = ''
        else:
            g.user_role = 'viewer'
            g.username = ''

    def require_role(*roles):
        """Decorator to restrict endpoint to specific roles."""
        def decorator(f):
            @functools.wraps(f)
            def wrapper(*args, **kwargs):
                if g.user_role not in roles:
                    return jsonify({'error': 'forbidden'}), 403
                return f(*args, **kwargs)
            return wrapper
        return decorator

    # ── Static serving ──────────────────────────────────────────────

    @app.route('/')
    def index():
        return send_from_directory(app.static_folder, 'index.html')

    # ── Auth endpoints ──────────────────────────────────────────────

    @app.route('/api/auth/login', methods=['POST'])
    def auth_login():
        data = request.get_json()
        role = data.get('role')
        username = data.get('username', '')
        password = data.get('password', '')

        if role not in ('data_entry', 'warehouse'):
            return jsonify({'error': 'invalid role'}), 400

        if not validate_credentials(role, username, password):
            return jsonify({'error': 'invalid credentials'}), 401

        ip = request.remote_addr
        if check_session_conflict(role, ip):
            return jsonify({'error': f'{role} role is already in use from another device'}), 409

        set_active_session(role, ip, username)
        session['role'] = role
        session['username'] = username
        return jsonify({'role': role, 'username': username, 'authenticated': True})

    @app.route('/api/auth/logout', methods=['POST'])
    def auth_logout():
        role = session.get('role')
        if role:
            clear_active_session(role)
        session.clear()
        return jsonify({'success': True})

    @app.route('/api/auth/session', methods=['GET'])
    def auth_session():
        return jsonify({
            'role': g.user_role,
            'username': getattr(g, 'username', ''),
            'authenticated': g.user_role != 'viewer',
        })

    @app.route('/api/item_codes', methods=['GET'])
    def get_item_codes():
        """Get the item codes as {code: description} dict."""
        item_codes_path = os.path.join(LIBRARY_ROOT, '.library', 'item_codes.json')
        if not os.path.exists(item_codes_path):
            return jsonify({})

        with open(item_codes_path, 'r') as f:
            item_codes = json.load(f)
        return jsonify(item_codes)

    @app.route('/api/item_codes', methods=['POST'])
    @require_role('data_entry')
    def add_item_code():
        """Add a new item code to .library/item_codes.json."""
        data = request.get_json()
        code = data.get('code', '').strip()
        description = data.get('description', '').strip()

        if not code:
            return jsonify({'error': 'code is required'}), 400
        if not description:
            return jsonify({'error': 'description is required'}), 400

        item_codes_path = os.path.join(LIBRARY_ROOT, '.library', 'item_codes.json')
        os.makedirs(os.path.dirname(item_codes_path), exist_ok=True)

        if os.path.exists(item_codes_path):
            with open(item_codes_path, 'r') as f:
                item_codes = json.load(f)
        else:
            item_codes = {}

        if code in item_codes:
            return jsonify({'error': f'code {code} already exists'}), 409

        item_codes[code] = description

        # Sort by key and write back
        sorted_codes = dict(sorted(item_codes.items()))
        with open(item_codes_path, 'w') as f:
            json.dump(sorted_codes, f, indent=4)

        return jsonify(sorted_codes)

    @app.route('/api/folders/mtime', methods=['GET'])
    def get_image_root_mtime():
        """Return the modification time of IMAGE_ROOT for change detection."""
        try:
            mtime = os.path.getmtime(IMAGE_ROOT)
        except OSError:
            mtime = 0
        return jsonify({'mtime': mtime})

    @app.route('/api/events', methods=['GET'])
    def sse_stream():
        """Server-Sent Events stream for filesystem change notifications.

        Emits:
          event: root_changed   — subfolder added/deleted under IMAGE_ROOT
          event: folder_changed — files changed inside a subfolder
                                  data: {"path": "<relative-folder-name>"}
          : keepalive           — comment sent every 25 s to prevent proxy close
        """
        q = subscribe()

        def generate():
            try:
                while True:
                    try:
                        event = q.get(timeout=25)
                        yield (
                            f"event: {event['type']}\n"
                            f"data: {json.dumps(event['data'])}\n\n"
                        )
                    except queue.Empty:
                        yield ": keepalive\n\n"
            except GeneratorExit:
                pass
            finally:
                unsubscribe(q)

        return Response(
            stream_with_context(generate()),
            mimetype='text/event-stream',
            headers={
                'Cache-Control': 'no-cache',
                'X-Accel-Buffering': 'no',  # disable nginx buffering
            },
        )

    # ── App settings endpoints ───────────────────────────────────────

    @app.route('/api/settings', methods=['GET'])
    def get_app_settings():
        """Return current app settings."""
        return jsonify({'image_root': IMAGE_ROOT})

    @app.route('/api/settings', methods=['POST'])
    @require_role('data_entry')
    def update_app_settings():
        """Update app settings (image_root, etc.)."""
        global IMAGE_ROOT
        data = request.get_json()
        new_root = (data.get('image_root') or '').strip()
        if not new_root:
            return jsonify({'error': 'image_root is required'}), 400
        if not os.path.isabs(new_root):
            return jsonify({'error': 'Path must be absolute (start with /)'}), 400
        if not os.path.isdir(new_root):
            return jsonify({'error': 'Directory not found or not accessible'}), 400
        settings = _load_settings()
        settings['image_root'] = new_root
        _save_settings(settings)
        IMAGE_ROOT = new_root
        return jsonify({'image_root': IMAGE_ROOT})

    # ── Folder endpoints ────────────────────────────────────────────

    @app.route('/api/folders', methods=['GET'])
    def list_folders():
        """List all imported folders, plus unimported directories.

        Detects folder renames by matching file inodes from disk to the DB.
        When a rename is detected, updates the folder's name/path and all
        image filepaths — thumbnails stay put since they're keyed by folder ID.
        """
        conn = get_connection()
        imported = rows_to_dicts(conn.execute(
            "SELECT * FROM folders ORDER BY name"
        ).fetchall())
        print(f"DEBUG: list_folders - Imported folders from DB: {imported}")
        sys.stdout.flush()

        # Build a set of inode->folder_id from DB for rename detection
        db_inodes = {}
        for f in imported:
            rows = conn.execute(
                "SELECT inode FROM images WHERE folder_id = ? AND status != 'deleted' LIMIT 5",
                (f['id'],)
            ).fetchall()
            for r in rows:
                if r['inode']:
                    db_inodes[r['inode']] = f['id']

        # Scan disk directories (flat + one level of nesting)
        disk_dirs = []
        groups = []
        # When IMAGE_ROOT is the project root, skip app infrastructure dirs
        _project_dirs = {'app', '__pycache__'} if IMAGE_ROOT == LIBRARY_ROOT else set()
        for entry in sorted(os.listdir(IMAGE_ROOT)):
            full_path = os.path.join(IMAGE_ROOT, entry)
            if not os.path.isdir(full_path) or entry.startswith('.') or entry in _project_dirs:
                continue
            image_files = [
                fn for fn in os.listdir(full_path)
                if fn.lower().endswith(('.jpg', '.jpeg', '.png', '.tiff', '.heic'))
                and not fn.startswith('.')
            ]
            if image_files:
                # Flat folder with images
                disk_dirs.append((entry, full_path, image_files, None))
            else:
                # Check for nested child folders with images
                child_count = 0
                for child_entry in sorted(os.listdir(full_path)):
                    child_path = os.path.join(full_path, child_entry)
                    if not os.path.isdir(child_path) or child_entry.startswith('.'):
                        continue
                    child_images = [
                        fn for fn in os.listdir(child_path)
                        if fn.lower().endswith(('.jpg', '.jpeg', '.png', '.tiff', '.heic'))
                        and not fn.startswith('.')
                    ]
                    if child_images:
                        rel_path = os.path.join(entry, child_entry)
                        disk_dirs.append((child_entry, child_path, child_images, entry))
                        child_count += 1
                if child_count > 0:
                    groups.append({'name': entry, 'folder_count': child_count})
        print(f"DEBUG: list_folders - Disk directories found: {[d[0] for d in disk_dirs]}")
        sys.stdout.flush()

        matched_ids = set()
        all_folders = []

        for entry, full_path, image_files, parent in disk_dirs:
            # Build the relative path (nested: "parent/child", flat: "entry")
            rel_path = os.path.join(parent, entry) if parent else entry

            # Direct path match
            match = next((f for f in imported if f['path'] == rel_path), None)

            # If no direct match, try inode matching (folder was renamed)
            if not match:
                sample_inodes = set()
                for fn in image_files[:5]:
                    try:
                        sample_inodes.add(os.stat(os.path.join(full_path, fn)).st_ino)
                    except OSError:
                        pass
                for ino in sample_inodes:
                    if ino in db_inodes:
                        fid = db_inodes[ino]
                        match = next((f for f in imported if f['id'] == fid), None)
                        if match:
                            # Update the folder name/path in DB
                            old_path = match['path']
                            conn.execute(
                                "UPDATE folders SET name = ?, path = ? WHERE id = ?",
                                (entry, rel_path, match['id'])
                            )
                            conn.execute(
                                "UPDATE images SET filepath = REPLACE(filepath, ?, ?) WHERE folder_id = ?",
                                (old_path + '/', rel_path + '/', match['id'])
                            )
                            conn.commit()
                            match['name'] = entry
                            match['path'] = rel_path
                            break

            if match:
                matched_ids.add(match['id'])
                match_dict = {**match, 'imported': True, 'parent': parent}
                if 'manual_reviewed' not in match_dict:
                    match_dict['manual_reviewed'] = False
                if 'warehouse_verified' not in match_dict:
                    match_dict['warehouse_verified'] = False
                all_folders.append(match_dict)
            else:
                all_folders.append({
                    'id': None,
                    'name': entry,
                    'path': rel_path,
                    'image_count': len(image_files),
                    'imported': False,
                    'manual_reviewed': False,
                    'warehouse_verified': False,
                    'parent': parent,
                })

        conn.close()

        # Viewer role: only sees manually reviewed, imported folders
        if g.user_role == 'viewer':
            all_folders = [f for f in all_folders if f.get('imported') and f.get('manual_reviewed')]

        print(f"DEBUG: list_folders - Final all_folders to return: {all_folders}")
        sys.stdout.flush()
        return jsonify({'folders': all_folders, 'groups': groups})

    @app.route('/api/folders/import', methods=['POST'])
    @require_role('data_entry')
    def import_folder():
        """Import a folder: scan, insert to DB, generate thumbnails."""
        from .importer import import_folder as do_import

        data = request.get_json()
        folder_path = data.get('path')
        if not folder_path:
            return jsonify({'error': 'path is required'}), 400

        full_path = os.path.join(IMAGE_ROOT, folder_path)
        if not os.path.isdir(full_path):
            return jsonify({'error': 'folder not found'}), 404

        try:
            result = do_import(LIBRARY_ROOT, folder_path, image_root=IMAGE_ROOT)
            return jsonify(result)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/folders/<int:folder_id>', methods=['GET'])
    def get_folder(folder_id):
        conn = get_connection()
        folder = row_to_dict(conn.execute(
            "SELECT *, manual_reviewed FROM folders WHERE id = ?", (folder_id,)
        ).fetchone())
        conn.close()
        if not folder:
            return jsonify({'error': 'not found'}), 404
        if 'manual_reviewed' not in folder:
            folder['manual_reviewed'] = False
        if 'warehouse_verified' not in folder:
            folder['warehouse_verified'] = False
        return jsonify(folder)

    @app.route('/api/folders/<int:folder_id>/manual-reviewed', methods=['PUT'])
    @require_role('data_entry')
    def update_manual_reviewed(folder_id):
        data = request.get_json()
        manual_reviewed = data.get('manual_reviewed')
        if manual_reviewed is None or not isinstance(manual_reviewed, bool):
            return jsonify({'error': 'manual_reviewed (boolean) is required'}), 400

        conn = get_connection()
        conn.execute(
            "UPDATE folders SET manual_reviewed = ? WHERE id = ?",
            (manual_reviewed, folder_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True})

    @app.route('/api/folders/<int:folder_id>/warehouse-verified', methods=['PUT'])
    @require_role('warehouse')
    def update_warehouse_verified(folder_id):
        data = request.get_json()
        warehouse_verified = data.get('warehouse_verified')
        if warehouse_verified is None or not isinstance(warehouse_verified, bool):
            return jsonify({'error': 'warehouse_verified (boolean) is required'}), 400

        conn = get_connection()
        conn.execute(
            "UPDATE folders SET warehouse_verified = ? WHERE id = ?",
            (warehouse_verified, folder_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True})

    # ── Image endpoints ─────────────────────────────────────────────

    @app.route('/api/folders/<int:folder_id>/images', methods=['GET'])
    def list_images(folder_id):
        # Sync folder with disk before returning images
        conn = get_connection()
        folder = conn.execute(
            "SELECT path FROM folders WHERE id = ?", (folder_id,)
        ).fetchone()
        conn.close()
        if folder:
            folder_path = folder['path']
            full_path = os.path.join(IMAGE_ROOT, folder_path)
            if os.path.isdir(full_path):
                try:
                    do_import(LIBRARY_ROOT, folder_path, image_root=IMAGE_ROOT)
                except Exception as e:
                    print(f"Sync error for folder {folder_path}: {e}")

        status_filter = request.args.get('status', 'active')
        conn = get_connection()
        if status_filter == 'all':
            images = rows_to_dicts(conn.execute(
                "SELECT * FROM images WHERE folder_id = ? ORDER BY sort_order",
                (folder_id,)
            ).fetchall())
        else:
            images = rows_to_dicts(conn.execute(
                "SELECT * FROM images WHERE folder_id = ? AND status = ? ORDER BY sort_order",
                (folder_id, status_filter)
            ).fetchall())
        conn.close()
        return jsonify(images)

    @app.route('/api/images/<int:image_id>', methods=['GET'])
    def get_image(image_id):
        conn = get_connection()
        image = row_to_dict(conn.execute(
            "SELECT * FROM images WHERE id = ?", (image_id,)
        ).fetchone())
        conn.close()
        if not image:
            return jsonify({'error': 'not found'}), 404
        return jsonify(image)

    @app.route('/api/images/<int:image_id>/thumbnail', methods=['GET'])
    def serve_thumbnail(image_id):
        conn = get_connection()
        image = row_to_dict(conn.execute(
            "SELECT * FROM images WHERE id = ?", (image_id,)
        ).fetchone())
        conn.close()
        if not image or not image['thumbnail_path']:
            return jsonify({'error': 'not found'}), 404

        thumb_path = os.path.join(LIBRARY_ROOT, '.library', 'thumbnails', image['thumbnail_path'])
        if not os.path.exists(thumb_path):
            return jsonify({'error': 'thumbnail file missing'}), 404
        return send_file(thumb_path, mimetype='image/jpeg')

    @app.route('/api/images/<int:image_id>/full', methods=['GET'])
    def serve_full_image(image_id):
        conn = get_connection()
        image = row_to_dict(conn.execute(
            "SELECT * FROM images WHERE id = ?", (image_id,)
        ).fetchone())
        conn.close()
        if not image:
            return jsonify({'error': 'not found'}), 404

        full_path = os.path.join(IMAGE_ROOT, image['filepath'])
        if not os.path.exists(full_path):
            return jsonify({'error': 'file missing'}), 404
        return send_file(full_path)

    @app.route('/api/images/<int:image_id>/rotate', methods=['POST'])
    @require_role('data_entry')
    def rotate_image(image_id):
        """Rotate an image 90 degrees clockwise on disk (full image + thumbnail)."""
        from PIL import Image as PILImage

        conn = get_connection()
        image = row_to_dict(conn.execute(
            "SELECT * FROM images WHERE id = ?", (image_id,)
        ).fetchone())
        if not image:
            conn.close()
            return jsonify({'error': 'not found'}), 404

        full_path = os.path.join(IMAGE_ROOT, image['filepath'])
        if not os.path.exists(full_path):
            conn.close()
            return jsonify({'error': 'file missing'}), 404

        try:
            from PIL import ImageOps

            # Rotate full image — normalize EXIF orientation first so
            # the browser and pixel data agree, then rotate.
            with PILImage.open(full_path) as img:
                img = ImageOps.exif_transpose(img)
                rotated = img.transpose(PILImage.Transpose.ROTATE_270)
                rotated.save(full_path, quality=95)
                new_width, new_height = rotated.size

            # Rotate thumbnail if it exists
            if image['thumbnail_path']:
                thumb_path = os.path.join(LIBRARY_ROOT, '.library', 'thumbnails', image['thumbnail_path'])
                if os.path.exists(thumb_path):
                    with PILImage.open(thumb_path) as thumb:
                        thumb = ImageOps.exif_transpose(thumb)
                        rotated_thumb = thumb.transpose(PILImage.Transpose.ROTATE_270)
                        rotated_thumb.save(thumb_path, quality=85)

            # Update dimensions in DB
            conn.execute(
                "UPDATE images SET width = ?, height = ?, modified_at = CURRENT_TIMESTAMP WHERE id = ?",
                (new_width, new_height, image_id)
            )
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'width': new_width, 'height': new_height})

        except Exception as e:
            conn.close()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/images/<int:image_id>/rename', methods=['PUT'])
    @require_role('data_entry')
    def rename_image(image_id):
        data = request.get_json()
        new_name = data.get('filename')
        if not new_name:
            return jsonify({'error': 'filename required'}), 400

        conn = get_connection()
        image = row_to_dict(conn.execute(
            "SELECT * FROM images WHERE id = ?", (image_id,)
        ).fetchone())
        if not image:
            conn.close()
            return jsonify({'error': 'not found'}), 404

        old_full = os.path.join(IMAGE_ROOT, image['filepath'])
        folder_path = os.path.dirname(image['filepath'])
        new_filepath = os.path.join(folder_path, new_name)
        new_full = os.path.join(IMAGE_ROOT, new_filepath)

        if os.path.exists(new_full):
            conn.close()
            return jsonify({'error': 'file already exists'}), 409

        os.rename(old_full, new_full)
        conn.execute(
            "UPDATE images SET filename = ?, filepath = ?, modified_at = CURRENT_TIMESTAMP WHERE id = ?",
            (new_name, new_filepath, image_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'filepath': new_filepath})

    @app.route('/api/images/<int:image_id>/status', methods=['PUT'])
    @require_role('data_entry')
    def update_image_status(image_id):
        from .actions import sync_actions_file

        data = request.get_json()
        new_status = data.get('status')
        if new_status not in ('active', 'marked_delete', 'marked_ocr'):
            return jsonify({'error': 'invalid status'}), 400

        conn = get_connection()
        conn.execute(
            "UPDATE images SET status = ?, modified_at = CURRENT_TIMESTAMP WHERE id = ?",
            (new_status, image_id)
        )
        conn.commit()
        conn.close()

        sync_actions_file(LIBRARY_ROOT)
        return jsonify({'success': True})

    @app.route('/api/images/bulk-status', methods=['POST'])
    @require_role('data_entry')
    def bulk_update_status():
        from .actions import sync_actions_file

        data = request.get_json()
        image_ids = data.get('image_ids', [])
        new_status = data.get('status')
        if new_status not in ('active', 'marked_delete', 'marked_ocr'):
            return jsonify({'error': 'invalid status'}), 400
        if not image_ids:
            return jsonify({'error': 'no image_ids provided'}), 400

        conn = get_connection()
        placeholders = ','.join('?' for _ in image_ids)
        conn.execute(
            f"UPDATE images SET status = ?, modified_at = CURRENT_TIMESTAMP WHERE id IN ({placeholders})",
            [new_status] + image_ids
        )
        conn.commit()
        conn.close()

        sync_actions_file(LIBRARY_ROOT)
        return jsonify({'success': True, 'updated': len(image_ids)})

    # ── Action endpoints ────────────────────────────────────────────

    @app.route('/api/actions/pending', methods=['GET'])
    def get_pending_actions():
        from .actions import read_actions_file
        actions = read_actions_file(LIBRARY_ROOT)
        return jsonify(actions)

    @app.route('/api/actions/execute', methods=['POST'])
    @require_role('data_entry')
    def execute_actions():
        from .actions import execute_pending_actions
        result = execute_pending_actions(LIBRARY_ROOT, image_root=IMAGE_ROOT)
        return jsonify(result)

    # ── Culling endpoints ───────────────────────────────────────────

    @app.route('/api/culling/start', methods=['POST'])
    @require_role('data_entry')
    def start_culling():
        data = request.get_json()
        image_ids = data.get('image_ids', [])
        folder_id = data.get('folder_id')
        if not image_ids or len(image_ids) < 2:
            return jsonify({'error': 'need at least 2 images'}), 400
        if not folder_id:
            return jsonify({'error': 'folder_id required'}), 400

        conn = get_connection()
        cur = conn.execute(
            "INSERT INTO culling_sessions (folder_id, image_ids, picked_image_id) VALUES (?, ?, ?)",
            (folder_id, json.dumps(image_ids), image_ids[0])
        )
        session_id = cur.lastrowid
        conn.commit()

        session = row_to_dict(conn.execute(
            "SELECT * FROM culling_sessions WHERE id = ?", (session_id,)
        ).fetchone())
        conn.close()
        session['image_ids'] = json.loads(session['image_ids'])
        return jsonify(session)

    @app.route('/api/culling/<int:session_id>', methods=['GET'])
    def get_culling_session(session_id):
        conn = get_connection()
        session = row_to_dict(conn.execute(
            "SELECT * FROM culling_sessions WHERE id = ?", (session_id,)
        ).fetchone())
        conn.close()
        if not session:
            return jsonify({'error': 'not found'}), 404
        session['image_ids'] = json.loads(session['image_ids'])
        return jsonify(session)

    @app.route('/api/culling/<int:session_id>/pick', methods=['PUT'])
    @require_role('data_entry')
    def pick_culling_image(session_id):
        data = request.get_json()
        picked_id = data.get('image_id')

        conn = get_connection()
        session = row_to_dict(conn.execute(
            "SELECT * FROM culling_sessions WHERE id = ?", (session_id,)
        ).fetchone())
        if not session:
            conn.close()
            return jsonify({'error': 'not found'}), 404

        image_ids = json.loads(session['image_ids'])
        if picked_id not in image_ids:
            conn.close()
            return jsonify({'error': 'image not in session'}), 400

        conn.execute(
            "UPDATE culling_sessions SET picked_image_id = ? WHERE id = ?",
            (picked_id, session_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'picked_image_id': picked_id})

    # ── Folder unit endpoint ────────────────────────────────────────

    @app.route('/api/folders/<int:folder_id>/unit', methods=['PUT'])
    @require_role('data_entry')
    def set_folder_unit(folder_id):
        data = request.get_json()
        weight_unit = data.get('weight_unit')
        if weight_unit not in ('kg', 'lbs'):
            return jsonify({'error': 'weight_unit must be kg or lbs'}), 400

        conn = get_connection()
        folder = conn.execute("SELECT id FROM folders WHERE id = ?", (folder_id,)).fetchone()
        if not folder:
            conn.close()
            return jsonify({'error': 'not found'}), 404

        conn.execute("UPDATE folders SET weight_unit = ? WHERE id = ?", (weight_unit, folder_id))
        conn.commit()
        conn.close()
        return jsonify({'weight_unit': weight_unit})

    # ── Folder ROI endpoint ──────────────────────────────────────

    @app.route('/api/folders/<int:folder_id>/roi', methods=['PUT'])
    @require_role('data_entry')
    def set_folder_roi(folder_id):
        data = request.get_json()
        cells = data.get('cells')
        if cells is None:
            return jsonify({'error': 'cells required'}), 400

        conn = get_connection()
        folder = conn.execute("SELECT id FROM folders WHERE id = ?", (folder_id,)).fetchone()
        if not folder:
            conn.close()
            return jsonify({'error': 'not found'}), 404

        roi_json = json.dumps(cells) if cells else None
        conn.execute("UPDATE folders SET ocr_roi = ? WHERE id = ?", (roi_json, folder_id))
        conn.commit()
        conn.close()
        return jsonify({'ocr_roi': roi_json})

    # ── OCR endpoints ────────────────────────────────────────────────

    @app.route('/api/ocr/process/<int:image_id>', methods=['POST'])
    @require_role('data_entry')
    def process_ocr_image(image_id):
        """Process a single image for OCR and return the result immediately."""
        from .ocr import process_image, save_ocr_result

        conn = get_connection()
        image = row_to_dict(conn.execute(
            "SELECT * FROM images WHERE id = ?", (image_id,)
        ).fetchone())
        if not image:
            conn.close()
            return jsonify({'error': 'image not found'}), 404

        # Look up folder's weight_unit and ocr_roi
        folder = row_to_dict(conn.execute(
            "SELECT weight_unit, ocr_roi FROM folders WHERE id = ?", (image['folder_id'],)
        ).fetchone())
        conn.close()
        weight_unit = (folder or {}).get('weight_unit') or 'kg'
        ocr_roi = None
        if folder and folder.get('ocr_roi'):
            try:
                ocr_roi = json.loads(folder['ocr_roi'])
            except (json.JSONDecodeError, TypeError):
                pass

        debug_dir = os.path.join(LIBRARY_ROOT, '.library', 'ocr_debug')

        try:
            result = process_image(LIBRARY_ROOT, image_id, debug_dir=debug_dir, weight_unit=weight_unit, ocr_roi=ocr_roi, image_root=IMAGE_ROOT)
            save_ocr_result(image_id, result)
            resp = {
                'image_id': image_id,
                'filename': result.get('renamed') or image['filename'],
                'tag': result['tag'],
                'scale_weight': result['scale_weight'],
                'handwritten_weight': result['handwritten_weight'],
                'status': result['status'],
                'error_message': result.get('error_message'),
                'pipeline': result.get('pipeline'),
            }
            if result.get('renamed'):
                resp['renamed'] = result['renamed']
            return jsonify(resp)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/ocr/vlm-status', methods=['GET'])
    def vlm_status():
        """Check if Ollama VLM is available."""
        try:
            import requests as req
            r = req.get('http://localhost:11434/api/tags', timeout=5)
            r.raise_for_status()
            models = [m['name'] for m in r.json().get('models', [])]
            has_model = any('minicpm-v' in m or 'minicpm-v2.6' in m for m in models)
            return jsonify({'available': has_model, 'models': models})
        except Exception:
            return jsonify({'available': False, 'models': []})

    @app.route('/api/ocr/submit', methods=['POST'])
    @require_role('data_entry')
    def submit_ocr():
        """Submit images for OCR processing.

        Expects JSON: { image_ids: [1, 2, 3] }
        Creates pending ocr_results rows, then processes in-request.
        """
        from .ocr import process_batch, save_ocr_result

        data = request.get_json()
        image_ids = data.get('image_ids', [])
        if not image_ids:
            return jsonify({'error': 'no image_ids provided'}), 400

        # Create pending rows for all images first
        conn = get_connection()
        for iid in image_ids:
            conn.execute(
                """INSERT INTO ocr_results (image_id, status)
                   VALUES (?, 'pending')
                   ON CONFLICT(image_id) DO UPDATE SET
                       status = 'pending', processed_at = NULL""",
                (iid,)
            )
        conn.commit()
        conn.close()

        try:
            summary = process_batch(LIBRARY_ROOT, image_ids, image_root=IMAGE_ROOT)
            return jsonify(summary)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/ocr/results/<int:folder_id>', methods=['GET'])
    def get_ocr_results(folder_id):
        """Get all OCR results for images in a folder."""
        conn = get_connection()
        results = rows_to_dicts(conn.execute(
            """SELECT o.*, i.filename, i.filepath, i.original_filename
               FROM ocr_results o
               JOIN images i ON o.image_id = i.id
               WHERE i.folder_id = ?
               ORDER BY i.sort_order""",
            (folder_id,)
        ).fetchall())
        conn.close()
        return jsonify(results)

    @app.route('/api/ocr/result/<int:image_id>', methods=['GET'])
    def get_ocr_result(image_id):
        """Get OCR result for a single image."""
        conn = get_connection()
        result = row_to_dict(conn.execute(
            """SELECT o.*, i.filename, i.filepath, i.original_filename
               FROM ocr_results o
               JOIN images i ON o.image_id = i.id
               WHERE o.image_id = ?""",
            (image_id,)
        ).fetchone())
        conn.close()
        if not result:
            return jsonify({'error': 'no OCR result'}), 404
        return jsonify(result)

    @app.route('/api/ocr/tag-roi/<int:image_id>', methods=['GET'])
    def get_tag_roi(image_id):
        """Serve the saved Tag ROI crop image for a given image."""
        conn = get_connection()
        img = row_to_dict(conn.execute(
            "SELECT filename FROM images WHERE id = ?", (image_id,)
        ).fetchone())
        conn.close()
        if not img:
            return jsonify({'error': 'image not found'}), 404
        base = os.path.splitext(img['filename'])[0]
        path = os.path.join(LIBRARY_ROOT, '.library', 'tag_crops', f'{base}_tagroi.jpg')
        if not os.path.exists(path):
            return jsonify({'error': 'no tag ROI crop'}), 404
        return send_file(path, mimetype='image/jpeg')

    @app.route('/api/ocr/led-crop/<int:image_id>', methods=['GET'])
    def get_led_crop(image_id):
        """Serve the saved LED crop image for a given image."""
        conn = get_connection()
        img = row_to_dict(conn.execute(
            "SELECT filename FROM images WHERE id = ?", (image_id,)
        ).fetchone())
        conn.close()
        if not img:
            return jsonify({'error': 'image not found'}), 404
        base = os.path.splitext(img['filename'])[0]
        led_path = os.path.join(LIBRARY_ROOT, '.library', 'led_crops', f'{base}_ledcrop.jpg')
        if not os.path.exists(led_path):
            return jsonify({'error': 'no LED crop'}), 404
        return send_file(led_path, mimetype='image/jpeg')

    @app.route('/api/ocr/result/<int:image_id>', methods=['PUT'])
    @require_role('data_entry', 'warehouse')
    def update_ocr_result(image_id):
        """Manually update OCR result (for human review corrections)."""
        import re
        evs_pattern = re.compile(r'^[A-Za-z]{3}\d{3}$')

        data = request.get_json()
        conn = get_connection()

        # Warehouse can only edit tare_weight
        allowed_keys = ('tag', 'scale_weight', 'handwritten_weight', 'status', 'item', 'tare_weight')
        if g.user_role == 'warehouse':
            allowed_keys = ('tare_weight',)

        # Enforce 30-char max on tag
        if 'tag' in data and data['tag']:
            data['tag'] = str(data['tag'])[:30]

        fields = []
        values = []
        for key in allowed_keys:
            if key in data:
                fields.append(f"{key} = ?")
                values.append(data[key])

        # If tag is being updated, check EVS validity
        new_tag = data.get('tag')
        if new_tag is not None and 'tag' in allowed_keys:
            is_evs = bool(new_tag) and evs_pattern.match(new_tag)
            if not is_evs:
                # Non-EVS: null out item, scale_weight, tare_weight
                for null_field in ('item', 'scale_weight', 'tare_weight'):
                    if f"{null_field} = ?" not in fields:
                        fields.append(f"{null_field} = ?")
                        values.append(None)
                    else:
                        # Override any value the client sent
                        idx = fields.index(f"{null_field} = ?")
                        values[idx] = None

        if not fields:
            conn.close()
            return jsonify({'error': 'no fields to update'}), 400

        # Ensure a row exists (warehouse may set tare_weight before OCR runs)
        conn.execute(
            "INSERT OR IGNORE INTO ocr_results (image_id, status) VALUES (?, 'pending')",
            (image_id,)
        )

        values.append(image_id)
        conn.execute(
            f"UPDATE ocr_results SET {', '.join(fields)} WHERE image_id = ?",
            values
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True})

    @app.route('/api/ocr/export/<int:folder_id>', methods=['GET'])
    def export_ocr_xlsx(folder_id):
        """Export OCR results for a folder as .xlsx.

        Template: Item | Tag | Gross | Tare | Net | Description
        - Tag and Gross come from ocr_results.
        - Net is a formula: Gross - Tare.
        - Item, Tare, Description are left blank.
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        conn = get_connection()
        folder = row_to_dict(conn.execute(
            "SELECT name FROM folders WHERE id = ?", (folder_id,)
        ).fetchone())
        if not folder:
            conn.close()
            return jsonify({'error': 'folder not found'}), 404

        import re
        tag_pattern = re.compile(r'^[A-Za-z]{3}\d{3}$')

        all_rows = rows_to_dicts(conn.execute(
            """SELECT o.tag, o.scale_weight, o.item, o.tare_weight
               FROM ocr_results o
               JOIN images i ON o.image_id = i.id
               WHERE i.folder_id = ?
               ORDER BY i.sort_order""",
            (folder_id,)
        ).fetchall())
        conn.close()

        # Only export rows with valid tags (3 letters + 3 digits)
        rows = [r for r in all_rows if r['tag'] and tag_pattern.match(r['tag'])]

        wb = Workbook()
        ws = wb.active
        ws.title = folder['name'][:31]  # sheet name max 31 chars

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 24

        # Header row
        headers = ['Item', 'Tag', 'Gross', 'Tare', 'Net', 'Description']
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            bottom=Side(style='thin', color='999999')
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Format column A (Item) as text
        from openpyxl.styles.numbers import FORMAT_TEXT
        for row_idx in range(2, len(rows) + 102):  # pre-format plenty of rows
            ws.cell(row=row_idx, column=1).number_format = FORMAT_TEXT

        # Data rows
        for i, r in enumerate(rows, 2):
            # A: Item
            if r['item']:
                ws.cell(row=i, column=1, value=r['item'])
            # B: Tag
            ws.cell(row=i, column=2, value=r['tag'])
            # C: Gross
            if r['scale_weight'] is not None:
                gross_cell = ws.cell(row=i, column=3, value=r['scale_weight'])
                gross_cell.number_format = '0.00'
            # D: Tare (from DB if available)
            if r.get('tare_weight') is not None:
                tare_cell = ws.cell(row=i, column=4, value=r['tare_weight'])
                tare_cell.number_format = '0.00'
            # E: Net = Gross - Tare (formula, 2 decimal places)
            net_cell = ws.cell(row=i, column=5, value=f'=C{i}-D{i}')
            net_cell.number_format = '0.00'
            # F: Description (blank)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"{folder['name']}_ocr_export.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    @app.route('/api/culling/<int:session_id>/finalize', methods=['POST'])
    @require_role('data_entry')
    def finalize_culling(session_id):
        from .actions import sync_actions_file

        conn = get_connection()
        session = row_to_dict(conn.execute(
            "SELECT * FROM culling_sessions WHERE id = ?", (session_id,)
        ).fetchone())
        if not session:
            conn.close()
            return jsonify({'error': 'not found'}), 404

        image_ids = json.loads(session['image_ids'])
        picked_id = session['picked_image_id']

        # Mark all non-picked images for deletion
        non_picked = [iid for iid in image_ids if iid != picked_id]
        if non_picked:
            placeholders = ','.join('?' for _ in non_picked)
            conn.execute(
                f"UPDATE images SET status = 'marked_delete', modified_at = CURRENT_TIMESTAMP WHERE id IN ({placeholders})",
                non_picked
            )

        conn.execute(
            "UPDATE culling_sessions SET status = 'finalized' WHERE id = ?",
            (session_id,)
        )
        conn.commit()
        conn.close()

        sync_actions_file(LIBRARY_ROOT)
        return jsonify({'success': True, 'marked_for_deletion': len(non_picked), 'picked_image_id': picked_id})

    return app
